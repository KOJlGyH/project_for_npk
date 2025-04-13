import telebot
from telebot import types
import requests
import json
import hashlib
import ast
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from io import BytesIO


bot = telebot.TeleBot('7635382318:AAEJS3sTRkzhQ32c86asRf7L-nQYX4o-fv8')
host = 'http://127.0.0.1:5000'
host = 'https://KOJlGylHl.pythonanywhere.com'


def convert_to_excel(data):
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active

    # Определяем стили
    header_font = Font(name='Times New Roman', size=14, bold=True)
    regular_font = Font(name='Times New Roman', size=14)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Стиль границ (только для нижней таблицы)
    table_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Цвета для оценок
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Определяем количество заданий
    num_tasks = len(data[0][1]) if data else 0

    # 1. Верхняя таблица (без границ)
    # Заголовки номеров заданий
    headers = ["Номера заданий"] + [str(i + 1) for i in range(num_tasks)] + ["Всего", "", ""]
    ws.append(headers)

    # Максимальные баллы за задания (без жирного шрифта)
    max_scores = [task[0] for task in data[0][1]] if data else []
    max_scores_row = ["Максимальный балл за задание"] + max_scores + [sum(max_scores), "", ""]
    ws.append(max_scores_row)

    # Пустая строка-разделитель между таблицами
    ws.append([])

    # 2. Нижняя таблица (с границами)
    # Заголовки для данных пользователей
    user_headers = ["Ученик", "Баллы за задания"] + [""] * (num_tasks - 1) + ["Сумма", "Решено, %",
                                                                              "Рекомендуемая оценка"]
    ws.append(user_headers)

    # Объединяем ячейки для "Баллы за задания"
    start_col = 2
    end_col = start_col + num_tasks - 1
    ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)

    # Данные пользователей
    for user_data in data:
        username = user_data[0]
        tasks = user_data[1]

        # Вычисляем баллы пользователя
        user_scores = [task[0] if task[1] else 0 for task in tasks]
        total_score = sum(user_scores)
        max_total = sum(max_scores)
        percentage = round((total_score / max_total) * 100, 2) if max_total != 0 else 0

        # Определяем рекомендуемую оценку
        if percentage >= 90:
            grade = 5
        elif percentage >= 75:
            grade = 4
        elif percentage >= 50:
            grade = 3
        elif percentage >= 25:
            grade = 2
        else:
            grade = 1

        # Формируем строку пользователя
        user_row = [username] + user_scores + [total_score, percentage, grade]
        ws.append(user_row)

        # Форматируем ячейку с процентами
        percent_cell = ws.cell(row=ws.max_row, column=len(user_row) - 1)
        percent_cell.number_format = FORMAT_NUMBER_00

        # Применяем цвет к ячейке с оценкой
        grade_cell = ws.cell(row=ws.max_row, column=len(user_row))
        if grade in (1, 2):
            grade_cell.fill = red_fill
        elif grade == 3:
            grade_cell.fill = yellow_fill
        elif grade == 4:
            grade_cell.fill = blue_fill
        elif grade == 5:
            grade_cell.fill = green_fill

    # Применяем стили:
    # 1. Для верхней таблицы (без границ)
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = regular_font
            cell.alignment = center_alignment
            cell.border = Border()  # Пустая граница (нет границ)

    # Заголовок "Номера заданий" делаем жирным
    for cell in ws[1]:
        cell.font = header_font

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.font = regular_font
            cell.alignment = center_alignment
            cell.border = table_border

        # Заголовок "Ученик" делаем жирным
    for cell in ws[4]:
        cell.font = header_font

        # Центрируем объединенную ячейку "Баллы за задания"
    merged_cell = ws.cell(row=4, column=2)
    merged_cell.alignment = center_alignment

    # Рассчитываем оптимальные ширины столбцов
    col_widths = {}

    # Определяем ширину для каждого столбца
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        col_widths[col] = (max_length + 2) * 1.2

    # Особые правила для столбцов с заданиями
    if num_tasks > 0:
        # Ширина объединенного заголовка "Баллы за задания"
        merged_header_width = len("Баллы за задания") * 1.5

        # Средняя ширина для столбцов с заданиями
        avg_task_width = merged_header_width / num_tasks

        # Определяем максимальную ширину среди столбцов с заданиями
        max_task_content_width = max(col_widths.get(col, 0) for col in range(2, 2 + num_tasks))

        # Выбираем наибольшую из средней ширины и ширины контента
        task_col_width = max(avg_task_width, max_task_content_width)

        # Устанавливаем одинаковую ширину для всех столбцов заданий
        for col in range(2, 2 + num_tasks):
            col_widths[col] = task_col_width

    # Устанавливаем рассчитанные ширины столбцов
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Сохраняем файл
    wb.save("результаты.xlsx")


def verify_password(password: str, salt: str, hashed: str) -> bool:
    new_hash = hashlib.pbkdf2_hmac(
        'sha256',
        password.encode('utf-8'),
        bytes.fromhex(salt),
        100000
    ).hex()
    return new_hash == hashed


def convert_to_excel(data):
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active

    # Определяем стили
    header_font = Font(name='Times New Roman', size=14, bold=True)
    regular_font = Font(name='Times New Roman', size=14)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Стиль границ (только для нижней таблицы)
    table_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Цвета для оценок
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    # Определяем количество заданий
    num_tasks = len(data[0][1]) if data else 0

    # 1. Верхняя таблица (без границ)
    # Заголовки номеров заданий
    headers = ["Номера заданий"] + [str(i + 1) for i in range(num_tasks)] + ["Всего", "", ""]
    ws.append(headers)

    # Максимальные баллы за задания (без жирного шрифта)
    max_scores = [task[0] for task in data[0][1]] if data else []
    max_scores_row = ["Максимальный балл за задание"] + max_scores + [sum(max_scores), "", ""]
    ws.append(max_scores_row)

    # Пустая строка-разделитель между таблицами
    ws.append([])

    # 2. Нижняя таблица (с границами)
    # Заголовки для данных пользователей
    user_headers = ["Ученик", "Баллы за задания"] + [""] * (num_tasks - 1) + ["Сумма", "Решено, %",
                                                                              "Рекомендуемая оценка"]
    ws.append(user_headers)

    # Объединяем ячейки для "Баллы за задания"
    start_col = 2
    end_col = start_col + num_tasks - 1
    ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)

    # Данные пользователей
    for user_data in data:
        username = user_data[0]
        tasks = user_data[1]

        # Вычисляем баллы пользователя
        user_scores = [task[0] if task[1] else 0 for task in tasks]
        total_score = sum(user_scores)
        max_total = sum(max_scores)
        percentage = round((total_score / max_total) * 100, 2) if max_total != 0 else 0

        # Определяем рекомендуемую оценку
        if percentage >= 90:
            grade = 5
        elif percentage >= 75:
            grade = 4
        elif percentage >= 50:
            grade = 3
        elif percentage >= 25:
            grade = 2
        else:
            grade = 1

        # Формируем строку пользователя
        user_row = [username] + user_scores + [total_score, percentage, grade]
        ws.append(user_row)

        # Форматируем ячейку с процентами
        percent_cell = ws.cell(row=ws.max_row, column=len(user_row) - 1)
        percent_cell.number_format = FORMAT_NUMBER_00

        # Применяем цвет к ячейке с оценкой
        grade_cell = ws.cell(row=ws.max_row, column=len(user_row))
        if grade in (1, 2):
            grade_cell.fill = red_fill
        elif grade == 3:
            grade_cell.fill = yellow_fill
        elif grade == 4:
            grade_cell.fill = blue_fill
        elif grade == 5:
            grade_cell.fill = green_fill

    # Применяем стили:
    # 1. Для верхней таблицы (без границ)
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = regular_font
            cell.alignment = center_alignment
            cell.border = Border()  # Пустая граница (нет границ)

    # Заголовок "Номера заданий" делаем жирным
    for cell in ws[1]:
        cell.font = header_font

    # 2. Для нижней таблицы (с границами)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.font = regular_font
            cell.alignment = center_alignment
            cell.border = table_border

    # Заголовок "Ученик" делаем жирным
    for cell in ws[4]:
        cell.font = header_font

    # Центрируем объединенную ячейку "Баллы за задания"
    merged_cell = ws.cell(row=4, column=2)
    merged_cell.alignment = center_alignment

    # Рассчитываем оптимальные ширины столбцов
    col_widths = {}

    # Определяем ширину для каждого столбца
    for col in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        col_widths[col] = (max_length + 2) * 1.2

    # Особые правила для столбцов с заданиями
    if num_tasks > 0:
        # Ширина объединенного заголовка "Баллы за задания"
        merged_header_width = len("Баллы за задания") * 1.5

        # Средняя ширина для столбцов с заданиями
        avg_task_width = merged_header_width / num_tasks

        # Определяем максимальную ширину среди столбцов с заданиями
        max_task_content_width = max(col_widths.get(col, 0) for col in range(2, 2 + num_tasks))

        # Выбираем наибольшую из средней ширины и ширины контента
        task_col_width = max(avg_task_width, max_task_content_width)

        # Устанавливаем одинаковую ширину для всех столбцов заданий
        for col in range(2, 2 + num_tasks):
            col_widths[col] = task_col_width

    # Устанавливаем рассчитанные ширины столбцов
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


    return wb


def get_from_db(data):
    data = json.dumps(data)
    return requests.get(host, data=data).json()


def post_in_db(data):
    data = json.dumps(data)
    requests.post(host, data=data)
    return None


def sign_up(message):
    login, password = message.text.split()
    print(login, password)
    # password = sha256_hash(password)
    logins = get_from_db(['login', 'user', 'id > 0'])
    logins = list(map(lambda x: x[0], logins))
    passwords = get_from_db(['password', 'user', f'login="{login}"'])
    hashed = list(map(lambda x: x[0], passwords))[0]
    salt = get_from_db(['salt', 'salts', f'login="{login}"'])
    salt = list(map(lambda x: x[0], salt))[0]
    print(passwords)

    markup = types.InlineKeyboardMarkup()
    reg = types.InlineKeyboardButton('Попробовать ещё', callback_data='chance')
    back_to_menu = types.InlineKeyboardButton('Вернуться в меню', callback_data='back')
    markup.add(reg)
    markup.add(back_to_menu)
    if login not in logins:
        bot.send_message(message.chat.id, 'Такой логин не зарегистрирован.', reply_markup=markup)
    elif not verify_password(password, salt, hashed):
        bot.send_message(message.chat.id, 'Неправильный пароль', reply_markup=markup)
    else:
        try:
            logins = get_from_db(['login', 'tg', f'true'])
            tg_ids = get_from_db(['tg_id', 'tg', 'true'])
            print(logins, login)
            if [message.chat.id] in tg_ids:
                post_in_db(['update', 'tg', 'tg_id="-1"', f'tg_id="{message.chat.id}"'])
            if [login] in logins:
                post_in_db(['update', 'tg', f'tg_id="{message.chat.id}"', f'login="{login}"'])
                bot.send_message(message.chat.id, 'Вы успешно зарегистрировались, уже начинаем продавать ваши персональные данные')
            else:
                post_in_db(['insert', 'tg(login, password, tg_id)', f'{login} {hashed} {message.chat.id}'])
                bot.send_message(message.chat.id, 'Вы успешно зарегистрировались, уже начинаем продавать ваши персональные данные')
        except:
            bot.send_message(message.chat.id, 'Упс, что-то пошло не так')


def results(message):
    key = message.text
    login = get_from_db(['login', 'tg', f'tg_id="{message.chat.id}"'])
    login = list(map(lambda x: x[0], login))[0]
    result = get_from_db(['results', 'test', f'author="{login}" and key="{key}"'])
    result = list(map(lambda x: x[0], result))
    if len(result) == 0:
        markup = types.InlineKeyboardMarkup()
        reg = types.InlineKeyboardButton('Отправить код ещё раз', callback_data='check_results')
        back_to_menu = types.InlineKeyboardButton('Вернуться в меню', callback_data='back')
        markup.add(reg)
        markup.add(back_to_menu)
        bot.send_message(message.chat.id, 'К сожалению я не нашёл такой работы', reply_markup=markup)
    else:
        result = result[0].rstrip('***').split('***')
        for i in range(len(result)):
            result[i] = result[i].replace('%', ',')
            log = result[i][:result[i].index('[')]
            res = ast.literal_eval(result[i][result[i].index('['):])
            print(log, res)
            result[i] = [log, res]
        print(result)
        for i in range(len(result)):
            login = result[i][0]
            username = get_from_db(['username', 'user', f'login="{login}"'])[0][0]
            result[i][0] = username
        print('Результат', result)
        wb = convert_to_excel(result)
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)  # Перемещаем указатель в начало файла

        # Отправляем файл
        bot.send_document(
            chat_id=message.chat.id,
            document=excel_file,
            caption="Ваш Excel-файл",
            visible_file_name="example.xlsx"
        )
        excel_file.close()




@bot.message_handler(commands=['start'])
def start_message(message):
    markup = types.InlineKeyboardMarkup()
    reg = types.InlineKeyboardButton('Регистрация', callback_data='registration')
    check_res = types.InlineKeyboardButton('Узнать результаты работы', callback_data='check_results')
    markup.add(reg)
    markup.add(check_res)
    bot.send_message(message.chat.id, 'Меню бота', reply_markup=markup)




@bot.callback_query_handler(func=lambda call: True)
def message_reply(call):
    if call.data == 'registration':
        m = bot.send_message(call.message.chat.id, 'Введие логин и пароль')
        bot.register_next_step_handler(m, sign_up)
    if call.data == 'check_results':
        m = bot.send_message(call.message.chat.id, 'Введите код доступа к вашей проверочной работе')
        bot.register_next_step_handler(m, results)
    if call.data == 'chance':
        m = bot.send_message(call.message.chat.id, 'И ещё одна попыточка)')
        bot.register_next_step_handler(m, sign_up)
    if call.data == 'back':
        start_message(call.message)



bot.infinity_polling()
